"""
Generate 3ioNetra Infrastructure Cost Excel — GCP Cost Calculator sheet.
Run: python3 generate_infra_cost_excel.py
Output: 3ioNetra_Infra_Cost.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Styles ──
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
SUB_FONT = Font(name="Calibri", bold=True, size=12, color="2B579A")
SUB_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2B579A")
NORMAL = Font(name="Calibri", size=11)
BOLD = Font(name="Calibri", bold=True, size=11)
TOTAL_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
TOTAL_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
RATE_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)


def write_row(ws, row, data, font=NORMAL, fill=None):
    for c, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = font
        if fill:
            cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def write_title(ws, row, title, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")


def write_sub(ws, row, title, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SUB_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def auto_width(ws, cols, min_w=14, max_w=42):
    from openpyxl.utils import get_column_letter
    for c in range(1, cols + 1):
        mx = min_w
        for row in ws.iter_rows(min_col=c, max_col=c):
            for cell in row:
                if cell.value:
                    mx = max(mx, min(len(str(cell.value)), max_w))
        ws.column_dimensions[get_column_letter(c)].width = mx + 3


# ── Workbook ──
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "GCP Cost Calculator"
COLS = 7
r = 1

# ================================================================
# TITLE
# ================================================================
write_title(ws, r, "3ioNetra — GCP Cost Calculator (Production)", COLS); r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
ws.cell(row=r, column=1, value="Exact pricing from GCP rate cards. Region: europe-west1 (Tier 1). All rates as of March 2026.").font = Font(name="Calibri", size=10, italic=True, color="666666")
r += 2

# ================================================================
# SECTION 1: GCP PRICING RATES
# ================================================================
write_sub(ws, r, "GCP Pricing Rates (Reference)", COLS); r += 1
headers = ["Service", "SKU", "Rate", "Unit", "Tier / Region", "", ""]
write_row(ws, r, headers, HEADER_FONT, HEADER_FILL); r += 1

rates = [
    ["Cloud Run", "CPU (always allocated)", "$0.00001800", "per vCPU-second", "Tier 1", "= $0.0648/vCPU-hr", ""],
    ["Cloud Run", "Memory (always allocated)", "$0.00000200", "per GiB-second", "Tier 1", "= $0.0072/GiB-hr", ""],
    ["Cloud Run", "Requests", "$0.40", "per 1 million", "All tiers", "", ""],
    ["Cloud Run", "Free tier — CPU", "180,000", "vCPU-seconds/mo", "All tiers", "= 50 vCPU-hours", ""],
    ["Cloud Run", "Free tier — Memory", "360,000", "GiB-seconds/mo", "All tiers", "= 100 GiB-hours", ""],
    ["Cloud Run", "Free tier — Requests", "2,000,000", "requests/mo", "All tiers", "", ""],
    ["Memorystore Redis", "Basic tier", "$0.049", "per GB per hour", "europe-west1", "", ""],
    ["VPC Connector", "e2-micro instance", "$0.0084", "per instance per hour", "europe-west1", "", ""],
    ["Artifact Registry", "Storage", "$0.10", "per GB per month", "All regions", "", ""],
    ["Cloud Storage", "Standard (regional)", "$0.020", "per GB per month", "europe-west1", "", ""],
    ["Cloud Storage", "Standard (multi-region US)", "$0.026", "per GB per month", "US", "", ""],
    ["Cloud Build", "Build minutes", "$0.003", "per build-minute", "After 120 free min/day", "", ""],
    ["Cloud Logging", "Log ingestion", "$0.50", "per GB", "After 50 GB free/mo", "", ""],
]
for row_data in rates:
    write_row(ws, r, row_data, NORMAL, RATE_FILL)
    r += 1

r += 1

# ================================================================
# SECTION 2: RESOURCE-BY-RESOURCE CALCULATION
# ================================================================
write_sub(ws, r, "Resource-by-Resource Cost Calculation", COLS); r += 1
headers = ["Resource", "SKU / Config", "Quantity", "Unit", "Rate ($/unit)", "Calculation", "Monthly Cost ($)"]
write_row(ws, r, headers, HEADER_FONT, HEADER_FILL); r += 1

# -- Cloud Run --
write_row(ws, r, ["CLOUD RUN — backend (min=1, 8 vCPU, 32 GiB)", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1

cr_rows = [
    ["  Min instance CPU", "CPU always allocated", "8 vCPU × 730 hrs", "vCPU-hours", "$0.0648", "8 × 730 × $0.0648", "$378.43"],
    ["  Min instance Memory", "Memory always allocated", "32 GiB × 730 hrs", "GiB-hours", "$0.0072", "32 × 730 × $0.0072", "$168.19"],
    ["  Free tier — CPU", "180,000 vCPU-sec", "50 vCPU-hrs", "vCPU-hours", "-$0.0648", "50 × $0.0648", "-$3.24"],
    ["  Free tier — Memory", "360,000 GiB-sec", "100 GiB-hrs", "GiB-hours", "-$0.0072", "100 × $0.0072", "-$0.72"],
    ["  Requests", "~500K requests/mo", "0", "(2M free)", "$0.40/M", "Within free tier", "$0.00"],
]
for row_data in cr_rows:
    write_row(ws, r, row_data)
    r += 1

write_row(ws, r, ["  CLOUD RUN SUBTOTAL (min instance only)", "", "", "", "", "", "$542.66"], BOLD, YELLOW_FILL); r += 1

r += 1
write_row(ws, r, ["CLOUD RUN — Burst (additional instances at peak)", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
burst_rows = [
    ["  Burst CPU (avg 1 extra inst, 20% time)", "CPU always allocated", "8 vCPU × 146 hrs", "vCPU-hours", "$0.0648", "8 × 146 × $0.0648", "$75.69"],
    ["  Burst Memory", "Memory always allocated", "32 GiB × 146 hrs", "GiB-hours", "$0.0072", "32 × 146 × $0.0072", "$33.64"],
]
for row_data in burst_rows:
    write_row(ws, r, row_data)
    r += 1
write_row(ws, r, ["  CLOUD RUN BURST SUBTOTAL", "", "", "", "", "", "$109.33"], BOLD, YELLOW_FILL); r += 1

r += 1

# -- Memorystore Redis --
write_row(ws, r, ["MEMORYSTORE REDIS — mitra-cache (Basic, 1 GB)", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
write_row(ws, r, ["  Redis Basic tier", "1 GB, 730 hrs/mo", "1 GB × 730 hrs", "GB-hours", "$0.049", "1 × 730 × $0.049", "$35.77"]); r += 1
write_row(ws, r, ["  REDIS SUBTOTAL", "", "", "", "", "", "$35.77"], BOLD, YELLOW_FILL); r += 1

r += 1

# -- VPC Connector --
write_row(ws, r, ["VPC CONNECTOR — mitra-connector (e2-micro, min=2)", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
write_row(ws, r, ["  e2-micro instances (min 2, always on)", "2 instances, 730 hrs/mo", "2 × 730 hrs", "instance-hours", "$0.0084", "2 × 730 × $0.0084", "$12.26"]); r += 1
write_row(ws, r, ["  VPC CONNECTOR SUBTOTAL", "", "", "", "", "", "$12.26"], BOLD, YELLOW_FILL); r += 1

r += 1

# -- Artifact Registry --
write_row(ws, r, ["ARTIFACT REGISTRY — cloud-run-source-deploy", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
write_row(ws, r, ["  Docker image storage", "~6.2 GB", "6.2 GB", "GB-months", "$0.10", "6.2 × $0.10", "$0.62"]); r += 1
write_row(ws, r, ["  ARTIFACT REGISTRY SUBTOTAL", "", "", "", "", "", "$0.62"], BOLD, YELLOW_FILL); r += 1

r += 1

# -- Cloud Storage --
write_row(ws, r, ["CLOUD STORAGE (3ioNetra buckets only)", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
gcs_rows = [
    ["  ionetra_cloudbuild", "US multi-region, 1.5 GB", "1.5 GB", "GB-months", "$0.026", "1.5 × $0.026", "$0.04"],
    ["  run-sources-ionetra-europe-west1", "europe-west1, 0.5 GB", "0.5 GB", "GB-months", "$0.020", "0.5 × $0.020", "$0.01"],
]
for row_data in gcs_rows:
    write_row(ws, r, row_data)
    r += 1
write_row(ws, r, ["  CLOUD STORAGE SUBTOTAL", "", "", "", "", "", "$0.05"], BOLD, YELLOW_FILL); r += 1

r += 1

# -- Cloud Build --
write_row(ws, r, ["CLOUD BUILD", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
write_row(ws, r, ["  Build minutes", "120 free min/day = 3,600/mo", "~100 min/mo used", "build-minutes", "$0.003", "Within free tier", "$0.00"]); r += 1
write_row(ws, r, ["  CLOUD BUILD SUBTOTAL", "", "", "", "", "", "$0.00"], BOLD, YELLOW_FILL); r += 1

r += 1

# -- Cloud Logging --
write_row(ws, r, ["CLOUD LOGGING", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
write_row(ws, r, ["  Log ingestion", "50 GB free/mo", "~5-10 GB/mo used", "GB", "$0.50", "Within free tier", "$0.00"]); r += 1
write_row(ws, r, ["  CLOUD LOGGING SUBTOTAL", "", "", "", "", "", "$0.00"], BOLD, YELLOW_FILL); r += 1

r += 2

# ================================================================
# SECTION 3: GCP TOTAL
# ================================================================
write_sub(ws, r, "GCP Cost Summary", COLS); r += 1
headers = ["Component", "", "", "", "", "Low ($/mo)", "High ($/mo)"]
write_row(ws, r, headers, HEADER_FONT, HEADER_FILL); r += 1

gcp_summary = [
    ["Cloud Run — min instance (always on)", "", "", "", "", "$542.66", "$542.66"],
    ["Cloud Run — burst (avg 1 extra, 20% time)", "", "", "", "", "$0.00", "$109.33"],
    ["Memorystore Redis (Basic 1 GB)", "", "", "", "", "$35.77", "$35.77"],
    ["VPC Connector (e2-micro × 2)", "", "", "", "", "$12.26", "$12.26"],
    ["Artifact Registry (6.2 GB)", "", "", "", "", "$0.62", "$0.62"],
    ["Cloud Storage (2 buckets, ~2 GB)", "", "", "", "", "$0.05", "$0.05"],
    ["Cloud Build", "", "", "", "", "$0.00", "$0.00"],
    ["Cloud Logging", "", "", "", "", "$0.00", "$0.00"],
]
for row_data in gcp_summary:
    write_row(ws, r, row_data)
    r += 1

write_row(ws, r, ["GCP SUBTOTAL", "", "", "", "", "$591.36", "$700.69"], TOTAL_FONT, TOTAL_FILL); r += 1

r += 2

# ================================================================
# SECTION 4: EXTERNAL SERVICES (not in GCP Calculator)
# ================================================================
write_sub(ws, r, "External Services (Not GCP)", COLS); r += 1
headers = ["Service", "Provider", "Tier / Plan", "Details", "", "Low ($/mo)", "High ($/mo)"]
write_row(ws, r, headers, HEADER_FONT, HEADER_FILL); r += 1

ext = [
    ["MongoDB Atlas", "MongoDB Inc.", "M10 dedicated", "2 vCPU, 2 GB RAM, 10 GB storage", "", "$57.00", "$57.00"],
    ["MongoDB Atlas (scale-up)", "MongoDB Inc.", "M30 (if needed)", "For 50K+ users", "", "$0.00", "$443.00"],
    ["Vercel (Frontend)", "Vercel", "Pro plan", "Next.js hosting, 1 TB bandwidth", "", "$20.00", "$20.00"],
    ["Gemini API", "Google", "Flash only (~10K convos)", "$0.003/convo × 10K", "", "$30.00", "$50.00"],
    ["Domain + DNS", "Cloudflare", "Free plan + domain renewal", "3iosetu.com, auto SSL", "", "$1.00", "$1.00"],
    ["Cloudflare Pro (optional)", "Cloudflare", "Pro plan", "WAF, advanced DDoS", "", "$0.00", "$20.00"],
]
for row_data in ext:
    write_row(ws, r, row_data)
    r += 1

write_row(ws, r, ["EXTERNAL SUBTOTAL", "", "", "", "", "$108.00", "$591.00"], TOTAL_FONT, TOTAL_FILL); r += 1

r += 2

# ================================================================
# SECTION 5: GRAND TOTAL
# ================================================================
write_sub(ws, r, "GRAND TOTAL — 3ioNetra Production Infrastructure", COLS); r += 1
headers = ["", "", "", "", "", "Low ($/mo)", "High ($/mo)"]
write_row(ws, r, headers, HEADER_FONT, HEADER_FILL); r += 1
write_row(ws, r, ["GCP Services", "", "", "", "", "$591.36", "$700.69"], BOLD); r += 1
write_row(ws, r, ["External Services", "", "", "", "", "$108.00", "$591.00"], BOLD); r += 1

r += 1
write_row(ws, r, ["TOTAL MONTHLY", "", "", "", "", "$699.36", "$1,291.69"], TOTAL_FONT, TOTAL_FILL); r += 1
write_row(ws, r, ["TOTAL ANNUAL", "", "", "", "", "$8,392.32", "$15,500.28"], BOLD, GREEN_FILL); r += 1

r += 2

# ================================================================
# SECTION 6: NOTES
# ================================================================
write_sub(ws, r, "Notes & Assumptions", COLS); r += 1
notes = [
    "1. Cloud Run rates are Tier 1 (europe-west1). CPU always-allocated pricing applies because min instances = 1.",
    "2. Burst estimate assumes avg 1 additional instance running 20% of the time (peak hours). Actual varies with traffic.",
    "3. Free tier deductions applied: 180K vCPU-sec, 360K GiB-sec, 2M requests, 120 build-min/day, 50 GB logs/mo.",
    "4. Memorystore Basic tier has no persistence or replication. Upgrade to Standard ($0.078/GB/hr) for HA.",
    "5. VPC Connector min=2 instances always running. Scales to max=10 under load (not billed unless active).",
    "6. MongoDB cost shows M10 ($57) as base. M30 ($443) only if scaling to 50K+ DAU.",
    "7. Gemini API uses Flash only (GEMINI_MODEL=gemini-2.5-flash). Pro is available but not enabled.",
    "8. All GCP values verified via gcloud CLI against live project 'ionetra' on 2026-03-26.",
]
for note in notes:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    cell = ws.cell(row=r, column=1, value=note)
    cell.font = Font(name="Calibri", size=10, color="555555")
    cell.alignment = Alignment(wrap_text=True)
    r += 1

auto_width(ws, COLS)

# ── Save ──
OUTPUT = "/Users/ankit1609/Desktop/3ioNetra/3ionetra/3ioNetra_Infra_Cost.xlsx"
wb.save(OUTPUT)
print(f"Generated: {OUTPUT}")
