"""
Generate 3ioNetra Infrastructure Cost Excel — AWS Equivalent of GCP Costing.
Mirrors the structure of GCP Costing.xlsx with AWS services and pricing.
Run: python3 generate_aws_cost_excel.py
Output: ~/Desktop/AWS_Costing.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles ──
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")  # AWS orange
SUB_FONT = Font(name="Calibri", bold=True, size=12, color="232F3E")  # AWS dark blue
SUB_FILL = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")  # Light orange
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="232F3E")
NORMAL = Font(name="Calibri", size=11)
BOLD = Font(name="Calibri", bold=True, size=11)
TOTAL_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
TOTAL_FILL = PatternFill(start_color="232F3E", end_color="232F3E", fill_type="solid")  # AWS dark
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
RATE_FILL = PatternFill(start_color="FEF3E8", end_color="FEF3E8", fill_type="solid")  # Light AWS orange
BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
ITALIC_GRAY = Font(name="Calibri", size=10, italic=True, color="666666")
NOTE_FONT = Font(name="Calibri", size=10, color="555555")


def write_row(ws, row, data, font=NORMAL, fill=None):
    for c, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = font
        if fill:
            cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def write_title(ws, row, title, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")


def write_sub(ws, row, title, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SUB_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def write_notes(ws, row, notes, cols):
    for note in notes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        cell = ws.cell(row=row, column=1, value=note)
        cell.font = NOTE_FONT
        cell.alignment = Alignment(wrap_text=True)
        row += 1
    return row


def auto_width(ws, cols, min_w=14, max_w=42):
    for c in range(1, cols + 1):
        mx = min_w
        for row in ws.iter_rows(min_col=c, max_col=c):
            for cell in row:
                if cell.value:
                    mx = max(mx, min(len(str(cell.value)), max_w))
        ws.column_dimensions[get_column_letter(c)].width = mx + 3


# ══════════════════════════════════════════════════════════════
# WORKBOOK
# ══════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════
# SHEET 1: Tech Stack
# ══════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Tech Stack"
COLS = 4
r = 1

write_title(ws, r, "3ioNetra - Tech Stack & Resource Requirements (AWS)", COLS); r += 2
write_sub(ws, r, "Tech Stack", COLS); r += 1
write_row(ws, r, ["Layer", "Technology", "Purpose", "Self-Hosted / Cloud"], HEADER_FONT, HEADER_FILL); r += 1

tech = [
    ["Backend", "FastAPI (Python 3.11)", "API server, ML inference, RAG pipeline", "Cloud (AWS Fargate ECS, eu-west-1)"],
    ["Frontend", "Next.js 14 (TypeScript)", "Chat UI, SSR", "Cloud (Vercel)"],
    ["Database", "MongoDB Atlas", "Users, conversations, auth tokens", "Cloud (MongoDB Atlas on AWS)"],
    ["Cache", "Redis 7.2", "Sessions, RAG cache, response cache", "Cloud (Amazon ElastiCache, eu-west-1)"],
    ["Vector DB", "Qdrant", "Scripture vector search", "Self-hosted (in backend)"],
    ["LLM", "Google Gemini 2.5 Flash / Pro", "Intent analysis, response generation", "Cloud API (Google)"],
    ["Embeddings", "multilingual-e5-large (1024-dim)", "Multilingual text embeddings", "Self-hosted (baked in Docker)"],
    ["Reranker", "BAAI/bge-reranker-v2-m3", "Neural reranking of search results", "Self-hosted (baked in Docker)"],
    ["Sparse Search", "SPLADE", "Keyword-level retrieval", "Self-hosted (baked in Docker)"],
    ["TTS", "gTTS", "Text-to-speech for verses", "Free API (Google Translate)"],
]
for row_data in tech:
    write_row(ws, r, row_data); r += 1

r += 1
write_sub(ws, r, "Resource Footprint", COLS); r += 1
write_row(ws, r, ["Resource", "Size", "Loaded When", "Notes"], HEADER_FONT, HEADER_FILL); r += 1

resources = [
    ["Embedding model (E5-large)", "1.3 GB", "Startup (eager)", "Loaded in RAM, required"],
    ["Reranker model (BGE-v2-m3)", "0.5 GB", "First search (lazy)", "Saves startup time"],
    ["SPLADE model", "0.8 GB", "First use (if enabled)", "Optional, improves keyword recall"],
    ["Embeddings data (embeddings.npy)", "377 MB", "Startup (memory-mapped)", "Zero heap cost, OS page cache"],
    ["Verse metadata (verses.json)", "108 MB", "Startup", "96K+ verses indexed"],
    ["Docker image (full)", "~3.0 GB", "Build time", "Models baked in, offline-capable"],
    ["Docker image (lite)", "~800 MB", "Build time", "Dev only, downloads models at runtime"],
    ["Per session (Redis)", "4-8 KB", "Per user", "60-min TTL, auto-purged"],
    ["Per conversation (MongoDB)", "15-20 KB", "Per conversation", "Permanent storage"],
    ["Minimum backend RAM", "3.5 GB", "-", "Models + verse index + OS overhead"],
]
for row_data in resources:
    write_row(ws, r, row_data); r += 1

auto_width(ws, COLS)

# ══════════════════════════════════════════════════════════════
# SHEET 2: Local - Development
# ══════════════════════════════════════════════════════════════
ws = wb.create_sheet("Local - Development")
COLS = 5
r = 1

write_title(ws, r, "Environment 1: Local / Development", COLS); r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
ws.cell(row=r, column=1, value="Developer laptop running everything via Docker Compose. No cloud costs.").font = ITALIC_GRAY; r += 2

write_sub(ws, r, "Hardware Requirements", COLS); r += 1
write_row(ws, r, ["Component", "Minimum", "Recommended", "Notes", "Cost"], HEADER_FONT, HEADER_FILL); r += 1
hw = [
    ["CPU", "4 cores", "8 cores", "For running 4 Docker containers", "$0"],
    ["RAM", "8 GB", "16 GB", "Backend alone needs 3.5 GB for ML models", "$0"],
    ["Storage", "10 GB free", "20 GB free", "Docker images + data + models", "$0"],
    ["GPU", "Not required", "Not required", "All inference is CPU-based", "$0"],
    ["OS", "macOS / Linux / WSL2", "macOS / Linux", "Docker Desktop required on macOS/Windows", "$0"],
]
for row_data in hw:
    write_row(ws, r, row_data); r += 1

r += 1
write_sub(ws, r, "Services (Docker Compose)", COLS); r += 1
write_row(ws, r, ["Service", "Docker Image", "Port", "Memory Usage", "Cost (USD/mo)"], HEADER_FONT, HEADER_FILL); r += 1
svcs = [
    ["Backend (FastAPI)", "Dockerfile.lite", "8080", "2-3 GB", "$0"],
    ["Frontend (Next.js)", "node:20-alpine (dev mode)", "3000", "200-400 MB", "$0"],
    ["Redis", "redis:alpine", "6379", "50-100 MB", "$0"],
    ["Qdrant", "qdrant/qdrant:latest", "6333, 6334", "500 MB - 1 GB", "$0"],
    ["MongoDB", "Local install or Atlas M0", "27017", "200-500 MB", "$0"],
]
for row_data in svcs:
    write_row(ws, r, row_data); r += 1

r += 1
write_sub(ws, r, "External Services", COLS); r += 1
write_row(ws, r, ["Service", "Tier", "Details", "Limit", "Cost (USD/mo)"], HEADER_FONT, HEADER_FILL); r += 1
ext_dev = [
    ["Gemini API", "Free tier", "For dev/testing", "15 RPM, 1M tokens/min", "$0"],
    ["MongoDB Atlas", "M0 (free)", "If using cloud instead of local", "512 MB storage", "$0"],
    ["Domain / SSL", "Not needed", "Using localhost", "-", "$0"],
    ["Monitoring", "Console logs", "stdout/stderr", "-", "$0"],
]
for row_data in ext_dev:
    write_row(ws, r, row_data); r += 1

r += 1
write_sub(ws, r, "Capacity", COLS); r += 1
write_row(ws, r, ["Metric", "Value", "", "", ""], HEADER_FONT, HEADER_FILL); r += 1
write_row(ws, r, ["Max concurrent sessions", "5-10", "", "", ""]); r += 1
write_row(ws, r, ["Suitable for", "1-2 developers testing", "", "", ""]); r += 1

r += 1
write_row(ws, r, ["Component", "Low", "High", "", ""], HEADER_FONT, HEADER_FILL); r += 1
write_row(ws, r, ["Gemini API (testing)", "$0", "$5", "", ""]); r += 1
write_row(ws, r, ["Everything else", "$0", "$0", "", ""]); r += 1
write_row(ws, r, ["TOTAL MONTHLY", "$0", "$5", "", ""], TOTAL_FONT, TOTAL_FILL); r += 1
write_row(ws, r, ["TOTAL ANNUAL", "$0", "$60", "", ""], BOLD, GREEN_FILL); r += 1

auto_width(ws, COLS)

# ══════════════════════════════════════════════════════════════
# SHEET 3: Staging - UAT
# ══════════════════════════════════════════════════════════════
ws = wb.create_sheet("Staging - UAT")
COLS = 5
r = 1

write_title(ws, r, "Environment 2: Staging / UAT (AWS)", COLS); r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
ws.cell(row=r, column=1, value="Lightweight AWS deployment for QA testing, client demos, and pre-production validation.").font = ITALIC_GRAY; r += 2

# User assumptions
write_sub(ws, r, "User Assumptions (This Pricing Is Based On)", COLS); r += 1
write_row(ws, r, ["Metric", "Value", "Notes", "", ""], HEADER_FONT, HEADER_FILL); r += 1
assumptions = [
    ["Concurrent users", "20 - 40", "Users actively chatting at the same time"],
    ["Daily active users (DAU)", "50 - 200", "Unique users per day"],
    ["Monthly active users (MAU)", "200 - 800", "Unique users per month"],
    ["Conversations per day", "100 - 500", "Total conversations across all users"],
    ["Conversations per month", "500 - 2,000", "Used for Gemini API cost calculation"],
    ["Avg turns per conversation", "6", "Each turn = 1 user message + 1 bot response"],
    ["Avg session duration", "5 - 10 minutes", "Per user per session"],
    ["Peak hour traffic", "10 - 20 concurrent", "Typically 2-3 hours during demos/testing"],
    ["Data storage growth", "~10 - 40 MB/month", "Conversations + user profiles"],
    ["Who is using this", "QA team, stakeholders, client demos", "Not real end users"],
]
for row_data in assumptions:
    write_row(ws, r, [row_data[0], row_data[1], row_data[2], "", ""]); r += 1

r += 1
# Backend Compute
write_sub(ws, r, "Backend Compute (AWS Fargate ECS)", COLS); r += 1
write_row(ws, r, ["Setting", "Value", "Notes", "", ""], HEADER_FONT, HEADER_FILL); r += 1
stg_settings = [
    ["Platform", "AWS Fargate (ECS)", "Managed containers"],
    ["Region", "eu-west-1 (Ireland)", "Low latency for EU / close to GCP equiv"],
    ["vCPU", "2", "Sufficient for light testing"],
    ["Memory", "4 GB", "Models (3.5 GB) + small headroom"],
    ["Desired count (min)", "0", "Scale to zero when idle (via scheduled scaling)"],
    ["Max tasks", "1", "Single task for UAT"],
    ["Docker image", "Full (models baked in)", "~3 GB image size"],
    ["Cold start", "15-20 seconds", "Fargate task launch + model loading"],
]
for row_data in stg_settings:
    write_row(ws, r, [row_data[0], row_data[1], row_data[2], "", ""]); r += 1

r += 1
write_row(ws, r, ["Cost Component", "Calculation", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
stg_compute = [
    ["vCPU", "2 vCPU x active hours x $0.04048/vCPU-hr", "$10", "$24", "~4-8 hrs active/day"],
    ["Memory", "4 GB x active hours x $0.004445/GB-hr", "$2", "$5", "Charged only when active"],
    ["ALB (shared or none)", "Optional for staging", "$0", "$16", "Can use public IP instead"],
    ["SUBTOTAL", "", "$12", "$45", ""],
]
for row_data in stg_compute:
    fill = YELLOW_FILL if row_data[0] == "SUBTOTAL" else None
    font = BOLD if row_data[0] == "SUBTOTAL" else NORMAL
    write_row(ws, r, row_data, font, fill); r += 1

r += 1
# Frontend
write_sub(ws, r, "Frontend (Vercel)", COLS); r += 1
write_row(ws, r, ["Setting", "Value", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
write_row(ws, r, ["Platform", "Vercel Free / Hobby", "$0", "$0", "100 GB bandwidth, unlimited deploys"]); r += 1

r += 1
# MongoDB
write_sub(ws, r, "Database - MongoDB Atlas", COLS); r += 1
write_row(ws, r, ["Setting", "Value", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
write_row(ws, r, ["Tier", "M0 (free) or M2 (shared)", "$0", "$9", "512 MB or 2 GB storage"]); r += 1
write_row(ws, r, ["Backups", "Daily (M2+)", "$0", "$0", "Included in M2"]); r += 1
write_row(ws, r, ["SUBTOTAL", "", "$0", "$9", ""], BOLD, YELLOW_FILL); r += 1

r += 1
# Redis
write_sub(ws, r, "Cache - Redis", COLS); r += 1
write_row(ws, r, ["Setting", "Value", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
write_row(ws, r, ["Provider", "ElastiCache cache.t4g.micro or Upstash", "$0", "$13", "Free tier: 750 hrs t3.micro (12 mo)"]); r += 1
write_row(ws, r, ["Memory", "0.5 - 1 GB", "$0", "$0", "Sufficient for UAT"]); r += 1
write_row(ws, r, ["SUBTOTAL", "", "$0", "$13", ""], BOLD, YELLOW_FILL); r += 1

r += 1
# Gemini API
write_sub(ws, r, "LLM - Gemini API", COLS); r += 1
write_row(ws, r, ["Setting", "Value", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
write_row(ws, r, ["Model", "Gemini 2.5 Flash", "-", "-", "Cheapest model for testing"]); r += 1
write_row(ws, r, ["Volume", "500-2K conversations/month", "-", "-", "QA + demo usage"]); r += 1
write_row(ws, r, ["Cost per conversation", "~$0.003", "-", "-", "6 turns avg"]); r += 1
write_row(ws, r, ["SUBTOTAL", "", "$2", "$10", ""], BOLD, YELLOW_FILL); r += 1

r += 1
# Networking & Other
write_sub(ws, r, "Networking, Storage & Other", COLS); r += 1
write_row(ws, r, ["Item", "Details", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
stg_net = [
    ["Domain (staging subdomain)", "staging.3iosetu.com", "$0", "$0", "Existing domain"],
    ["SSL certificate", "AWS Certificate Manager (free)", "$0", "$0", "Free with ALB"],
    ["DNS", "Cloudflare free tier", "$0", "$0", "Free"],
    ["NAT Gateway", "Not needed for staging (use public subnet)", "$0", "$0", "Save cost by using public IP"],
    ["ECR", "Docker images (<5 GB)", "$0", "$2", "Storage cost"],
    ["Monitoring", "CloudWatch basic", "$0", "$0", "Free tier"],
    ["SUBTOTAL", "", "$0", "$2", ""],
]
for row_data in stg_net:
    fill = YELLOW_FILL if row_data[0] == "SUBTOTAL" else None
    font = BOLD if row_data[0] == "SUBTOTAL" else NORMAL
    write_row(ws, r, row_data, font, fill); r += 1

r += 1
write_sub(ws, r, "Capacity", COLS); r += 1
write_row(ws, r, ["Metric", "Value", "", "", ""], HEADER_FONT, HEADER_FILL); r += 1
write_row(ws, r, ["Max concurrent sessions", "20-40", "", "", ""]); r += 1
write_row(ws, r, ["Max daily active users", "50-200", "", "", ""]); r += 1
write_row(ws, r, ["Suitable for", "QA team, client demos, UAT testing", "", "", ""]); r += 1

r += 1
write_sub(ws, r, "STAGING MONTHLY TOTAL", COLS); r += 1
write_row(ws, r, ["Component", "Low (USD/mo)", "High (USD/mo)", "", ""], HEADER_FONT, HEADER_FILL); r += 1
stg_total = [
    ["Backend (Fargate, 2 vCPU/4 GB, desired=0-1)", "$12", "$45"],
    ["Frontend (Vercel Free)", "$0", "$0"],
    ["MongoDB (Atlas M0/M2)", "$0", "$9"],
    ["Redis (ElastiCache free tier / Upstash)", "$0", "$13"],
    ["Gemini API (~1K convos)", "$2", "$10"],
    ["Docker Registry / Networking", "$0", "$2"],
    ["Domain / SSL / DNS", "$0", "$0"],
    ["Monitoring", "$0", "$0"],
]
for row_data in stg_total:
    write_row(ws, r, [row_data[0], row_data[1], row_data[2], "", ""]); r += 1

write_row(ws, r, ["TOTAL MONTHLY", "$14", "$79", "", ""], TOTAL_FONT, TOTAL_FILL); r += 1
write_row(ws, r, ["TOTAL ANNUAL", "$168", "$948", "", ""], BOLD, GREEN_FILL); r += 1

auto_width(ws, COLS)

# ══════════════════════════════════════════════════════════════
# SHEET 4: Production
# ══════════════════════════════════════════════════════════════
ws = wb.create_sheet("Production")
COLS = 5
r = 1

write_title(ws, r, "Environment 3: Production (AWS)", COLS); r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
ws.cell(row=r, column=1, value="Full AWS deployment for real users. Always warm, auto-scaling, monitoring, zero cold starts.").font = ITALIC_GRAY; r += 2

# User assumptions (same as GCP)
write_sub(ws, r, "User Assumptions (This Pricing Is Based On)", COLS); r += 1
write_row(ws, r, ["Metric", "Value", "Notes", "", ""], HEADER_FONT, HEADER_FILL); r += 1
prod_assumptions = [
    ["Concurrent users", "200 - 800", "Users actively chatting at the same time"],
    ["Daily active users (DAU)", "5,000 - 50,000", "Unique users per day"],
    ["Monthly active users (MAU)", "20,000 - 150,000", "Unique users per month"],
    ["Conversations per day", "10,000 - 50,000", "Total conversations across all users"],
    ["Conversations per month", "10,000 - 100,000", "Used for Gemini API cost calculation"],
    ["Avg turns per conversation", "6", "Each turn = 1 user message + 1 bot response"],
    ["Avg session duration", "5 - 10 minutes", "Per user per session"],
    ["Peak hour traffic", "100 - 400 concurrent", "Typically 3-4 peak hours/day"],
    ["Sessions per day (at 5K DAU)", "15,000 - 25,000", "Multiple sessions per user"],
    ["Data storage growth", "~500 MB - 2 GB/month", "Conversations + user profiles + feedback"],
    ["ECS tasks at peak", "3 - 100", "Auto-scaled based on traffic"],
    ["Who is using this", "Real end users (public)", "Full production traffic"],
]
for row_data in prod_assumptions:
    write_row(ws, r, [row_data[0], row_data[1], row_data[2], "", ""]); r += 1

r += 1
# Backend Compute
write_sub(ws, r, "Backend Compute (AWS Fargate ECS)", COLS); r += 1
write_row(ws, r, ["Setting", "Value", "Notes", "", ""], HEADER_FONT, HEADER_FILL); r += 1
prod_settings = [
    ["Platform", "AWS Fargate (ECS)", "Managed, auto-scaling"],
    ["Region", "eu-west-1 (Ireland)", "Equivalent to GCP europe-west1"],
    ["vCPU", "8", "Handles concurrent RAG + LLM calls"],
    ["Memory", "32 GB", "ML models + high headroom for burst traffic"],
    ["Desired count (min)", "1", "Always warm, no cold starts"],
    ["Max tasks", "100", "Scale for peak traffic via ECS auto-scaling"],
    ["Concurrency", "80 requests/task", "Max 8,000 concurrent at full scale"],
    ["Docker image", "Full (TRANSFORMERS_OFFLINE=1)", "~3 GB, no runtime downloads"],
    ["Workers", "1 (uvicorn async)", "Single-process, async concurrency"],
    ["Reranker", "Disabled (RERANKER_ENABLED=false)", "Saves ~0.5 GB RAM at runtime"],
    ["SPLADE", "Disabled (SPLADE_ENABLED=false)", "Saves ~0.8 GB RAM at runtime"],
]
for row_data in prod_settings:
    write_row(ws, r, [row_data[0], row_data[1], row_data[2], "", ""]); r += 1

r += 1
write_row(ws, r, ["Cost Component", "Calculation", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
prod_compute = [
    ["vCPU (min task, always on)", "8 vCPU x 730 hrs x $0.04048/hr", "$236.40", "$236.40", "Fixed cost (always warm)"],
    ["Memory (min task, always on)", "32 GB x 730 hrs x $0.004445/hr", "$103.82", "$103.82", "Fixed cost"],
    ["vCPU (burst tasks)", "8 vCPU x ~146 hrs x $0.04048/hr", "$0", "$47.28", "~20% of time at peak"],
    ["Memory (burst tasks)", "32 GB x ~146 hrs x $0.004445/hr", "$0", "$20.76", "Proportional to burst"],
    ["SUBTOTAL", "", "$340.22", "$408.26", ""],
]
for row_data in prod_compute:
    fill = YELLOW_FILL if row_data[0] == "SUBTOTAL" else None
    font = BOLD if row_data[0] == "SUBTOTAL" else NORMAL
    write_row(ws, r, row_data, font, fill); r += 1

r += 1
# ALB
write_sub(ws, r, "Application Load Balancer (ALB)", COLS); r += 1
write_row(ws, r, ["Setting", "Value", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
alb_rows = [
    ["ALB hourly charge", "730 hrs x $0.0225/hr", "$16.43", "$16.43", "Required for Fargate (Cloud Run has built-in LB)"],
    ["LCU charges", "~5-10 LCU-hours avg", "$5.00", "$10.00", "Based on connections, bandwidth, rules"],
    ["SUBTOTAL", "", "$21.43", "$26.43", ""],
]
for row_data in alb_rows:
    fill = YELLOW_FILL if row_data[0] == "SUBTOTAL" else None
    font = BOLD if row_data[0] == "SUBTOTAL" else NORMAL
    write_row(ws, r, row_data, font, fill); r += 1

r += 1
# Frontend
write_sub(ws, r, "Frontend (Vercel Pro)", COLS); r += 1
write_row(ws, r, ["Setting", "Value", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
write_row(ws, r, ["Vercel Pro", "1 TB bandwidth, analytics, team", "$20", "$20", "Custom domain included"]); r += 1

r += 1
# MongoDB
write_sub(ws, r, "Database - MongoDB Atlas", COLS); r += 1
write_row(ws, r, ["Setting", "Value", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
prod_mongo = [
    ["Tier", "M10 (dedicated)", "$57", "$57", "2 vCPU, 2 GB RAM"],
    ["Storage", "10 GB (auto-scales)", "$0", "$0", "Included in M10"],
    ["Backups", "Continuous + point-in-time recovery", "$0", "$0", "Included in M10"],
    ["Scale-up (50K+ users)", "M30 dedicated", "$0", "$443", "If needed"],
    ["SUBTOTAL", "", "$57", "$100", ""],
]
for row_data in prod_mongo:
    fill = YELLOW_FILL if row_data[0] == "SUBTOTAL" else None
    font = BOLD if row_data[0] == "SUBTOTAL" else NORMAL
    write_row(ws, r, row_data, font, fill); r += 1

r += 1
# ElastiCache Redis
write_sub(ws, r, "Cache - Amazon ElastiCache for Redis", COLS); r += 1
write_row(ws, r, ["Setting", "Value", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
prod_redis = [
    ["Provider", "Amazon ElastiCache for Redis", "-", "-", "Managed by AWS, VPC access"],
    ["Instance type", "cache.t4g.small (1.37 GiB)", "-", "-", "Closest to GCP 1 GB Memorystore"],
    ["Tier", "Single node (no replication)", "$20.44", "$20.44", "730 hrs x $0.028/hr"],
    ["Redis version", "7.2", "$0", "$0", "Latest stable"],
    ["Max connections", "20 (async pool)", "$0", "$0", "Configured in backend"],
    ["Cache: Sessions", "60-min TTL, ~500 KB per 100 sessions", "-", "-", "Auto-purged"],
    ["Cache: RAG results", "1-hour TTL, ~100 KB per entry", "-", "-", "Saves RAG compute"],
    ["Cache: Response", "6-hour TTL, ~200 KB per entry", "-", "-", "Semantic dedup"],
    ["Estimated total usage", "200-800 MB", "-", "-", "Well within 1.37 GiB"],
    ["Scale-up (if needed)", "cache.r6g.large (HA)", "$0", "$20", "For 50K+ users"],
    ["SUBTOTAL", "", "$20.44", "$40.44", ""],
]
for row_data in prod_redis:
    fill = YELLOW_FILL if row_data[0] == "SUBTOTAL" else None
    font = BOLD if row_data[0] == "SUBTOTAL" else NORMAL
    write_row(ws, r, row_data, font, fill); r += 1

r += 1
# Gemini API
write_sub(ws, r, "LLM - Gemini API", COLS); r += 1
write_row(ws, r, ["Setting", "Value", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
prod_gemini = [
    ["Current model (deployed)", "Gemini 2.5 Flash ONLY", "-", "-", "GEMINI_MODEL=gemini-2.5-flash"],
    ["Flash pricing", "$0.15 input / $0.60 output per 1M tokens", "-", "-", "All calls use Flash currently"],
    ["Volume estimate", "10K conversations/month", "-", "-", "~5K daily active users"],
    ["Cost per conversation (Flash)", "~$0.003", "-", "-", "6 turns avg"],
    ["Context caching savings", "~60-70% on input tokens", "-", "-", "System instruction cached 6h"],
    ["SUBTOTAL (Flash only)", "", "$30", "$50", ""],
]
for row_data in prod_gemini:
    fill = YELLOW_FILL if row_data[0].startswith("SUBTOTAL") else None
    font = BOLD if row_data[0].startswith("SUBTOTAL") else NORMAL
    write_row(ws, r, row_data, font, fill); r += 1

r += 1
# Networking
write_sub(ws, r, "Networking & Security", COLS); r += 1
write_row(ws, r, ["Item", "Details", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
prod_net = [
    ["Domain renewal", ".com annual renewal", "$1", "$1", "$12/year"],
    ["SSL certificate", "AWS Certificate Manager (free with ALB)", "$0", "$0", "Free"],
    ["Cloudflare DNS", "Free plan", "$0", "$0", "DNS + basic DDoS"],
    ["Cloudflare Pro (optional)", "WAF, analytics, advanced DDoS", "$0", "$20", "Recommended"],
    ["NAT Gateway", "For private subnet access to internet", "$35.04", "$37.44", "730 hrs x $0.048 + data processing"],
    ["AWS WAF (optional)", "Web application firewall", "$0", "$25", "If DDoS protection needed"],
    ["Egress (outbound traffic)", "~50 GB/mo", "$4.50", "$9.00", "$0.09/GB after first 100 GB free"],
    ["SUBTOTAL", "", "$40.54", "$92.44", ""],
]
for row_data in prod_net:
    fill = YELLOW_FILL if row_data[0] == "SUBTOTAL" else None
    font = BOLD if row_data[0] == "SUBTOTAL" else NORMAL
    write_row(ws, r, row_data, font, fill); r += 1

r += 1
# Storage & Docker Registry
write_sub(ws, r, "Storage & Docker Registry", COLS); r += 1
write_row(ws, r, ["Item", "Details", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
prod_storage = [
    ["Amazon ECR", "Docker images (~6.2 GB)", "$0.62", "$0.62", "$0.10/GB/month"],
    ["S3: CodeBuild artifacts", "Build artifacts (~1.5 GB)", "$0.03", "$0.03", "$0.023/GB/month"],
    ["S3: Source uploads", "Deployment sources (~0.5 GB)", "$0.01", "$0.01", "$0.023/GB/month"],
    ["CloudWatch Logs", "App logs via stdout", "$0", "$5", "5 GB free/month"],
    ["SUBTOTAL", "", "$0.66", "$5.66", ""],
]
for row_data in prod_storage:
    fill = YELLOW_FILL if row_data[0] == "SUBTOTAL" else None
    font = BOLD if row_data[0] == "SUBTOTAL" else NORMAL
    write_row(ws, r, row_data, font, fill); r += 1

r += 1
# Monitoring
write_sub(ws, r, "Monitoring, Logging & Alerts", COLS); r += 1
write_row(ws, r, ["Item", "Details", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
prod_mon = [
    ["CloudWatch Metrics", "CPU, memory, request metrics", "$0", "$0", "Basic metrics free"],
    ["CloudWatch Alarms", "Email/Slack on errors, latency", "$0", "$0", "10 free alarms"],
    ["CloudWatch Dashboards", "Custom dashboards", "$0", "$3", "$3/dashboard after 3 free"],
    ["X-Ray (optional)", "Distributed request tracing", "$0", "$10", "For debugging latency"],
    ["External monitoring (optional)", "UptimeRobot / Better Uptime", "$0", "$7", "Additional uptime alerts"],
    ["SUBTOTAL", "", "$0", "$20", ""],
]
for row_data in prod_mon:
    fill = YELLOW_FILL if row_data[0] == "SUBTOTAL" else None
    font = BOLD if row_data[0] == "SUBTOTAL" else NORMAL
    write_row(ws, r, row_data, font, fill); r += 1

r += 1
# Backup & DR
write_sub(ws, r, "Backup & Disaster Recovery", COLS); r += 1
write_row(ws, r, ["Item", "Details", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
prod_backup = [
    ["MongoDB Atlas backups", "Continuous + point-in-time", "$0", "$0", "Included in M10+"],
    ["ElastiCache persistence", "Single node (no persistence)", "$0", "$0", "Sessions are ephemeral"],
    ["Docker image versioning", "3 recent versions in ECR", "$0", "$0", "Included in ECR cost"],
    ["Scripture data", "Static, baked in image, git versioned", "$0", "$0", "No additional backup"],
    ["Manual DB export", "S3 bucket", "$0", "$2", "Monthly export"],
    ["SUBTOTAL", "", "$0", "$2", ""],
]
for row_data in prod_backup:
    fill = YELLOW_FILL if row_data[0] == "SUBTOTAL" else None
    font = BOLD if row_data[0] == "SUBTOTAL" else NORMAL
    write_row(ws, r, row_data, font, fill); r += 1

r += 1
# CI/CD
write_sub(ws, r, "CI/CD Pipeline", COLS); r += 1
write_row(ws, r, ["Item", "Details", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
prod_cicd = [
    ["AWS CodeBuild", "Triggered on deploy", "$0", "$5", "100 free min/month"],
    ["Vercel builds", "Auto-deploy on git push", "$0", "$0", "Included in Vercel"],
    ["GitHub Actions (optional)", "Tests, linting", "$0", "$0", "2000 free min/mo"],
    ["SUBTOTAL", "", "$0", "$5", ""],
]
for row_data in prod_cicd:
    fill = YELLOW_FILL if row_data[0] == "SUBTOTAL" else None
    font = BOLD if row_data[0] == "SUBTOTAL" else NORMAL
    write_row(ws, r, row_data, font, fill); r += 1

r += 1
# Secrets
write_sub(ws, r, "Secret Management", COLS); r += 1
write_row(ws, r, ["Item", "Details", "Low (USD/mo)", "High (USD/mo)", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
write_row(ws, r, ["AWS Secrets Manager", "4 secrets (Gemini, MongoDB, Redis, DB password)", "$1.60", "$1.60", "$0.40/secret/month"]); r += 1

r += 1
# Capacity
write_sub(ws, r, "Capacity", COLS); r += 1
write_row(ws, r, ["Metric", "Value", "Notes", "", ""], HEADER_FONT, HEADER_FILL); r += 1
capacity = [
    ["Concurrency per task", "80 requests", "Configured in ECS task definition"],
    ["Min tasks (always warm)", "1", "No cold starts"],
    ["Max tasks", "100", "Auto-scales with ECS auto-scaling"],
    ["Max concurrent sessions", "8,000", "100 tasks x 80"],
    ["Average response time", "2-4 seconds", "Including LLM streaming"],
    ["Effective throughput", "20-40 requests/second", "Sustained"],
    ["Daily active users supported", "5,000 - 50,000", "Depends on usage pattern"],
]
for row_data in capacity:
    write_row(ws, r, [row_data[0], row_data[1], row_data[2], "", ""]); r += 1

r += 1
# PRODUCTION TOTAL
write_sub(ws, r, "PRODUCTION MONTHLY TOTAL", COLS); r += 1
write_row(ws, r, ["Component", "Low (USD/mo)", "High (USD/mo)", "", ""], HEADER_FONT, HEADER_FILL); r += 1
prod_total = [
    ["Backend (Fargate, 8 vCPU/32 GB, min=1, max=100)", "$340.22", "$408.26"],
    ["Application Load Balancer", "$21.43", "$26.43"],
    ["Frontend (Vercel Pro)", "$20", "$20"],
    ["MongoDB Atlas (M10 dedicated)", "$57", "$100"],
    ["Redis (ElastiCache cache.t4g.small)", "$20.44", "$40.44"],
    ["Gemini API (~10K convos, Flash only)", "$30", "$50"],
    ["Networking (NAT GW, DNS, egress, Cloudflare)", "$40.54", "$92.44"],
    ["Storage (ECR + S3 + CloudWatch Logs)", "$0.66", "$5.66"],
    ["Monitoring, Logging & Alerts", "$0", "$20"],
    ["Backup & Disaster Recovery", "$0", "$2"],
    ["CI/CD Pipeline (CodeBuild)", "$0", "$5"],
    ["Secret Management", "$1.60", "$1.60"],
]
for row_data in prod_total:
    write_row(ws, r, [row_data[0], row_data[1], row_data[2], "", ""]); r += 1

write_row(ws, r, ["TOTAL MONTHLY", "$531.89", "$771.83", "", ""], TOTAL_FONT, TOTAL_FILL); r += 1
write_row(ws, r, ["TOTAL ANNUAL", "$6,382.68", "$9,261.96", "", ""], BOLD, GREEN_FILL); r += 1

auto_width(ws, COLS)

# ══════════════════════════════════════════════════════════════
# SHEET 5: Gemini API Cost (identical to GCP — external service)
# ══════════════════════════════════════════════════════════════
ws = wb.create_sheet("Gemini API Cost")
COLS = 5
r = 1

write_title(ws, r, "Gemini API - Detailed Cost Breakdown", COLS); r += 2

write_sub(ws, r, "Pricing (per 1 Million Tokens)", COLS); r += 1
write_row(ws, r, ["Model", "Input Tokens", "Output Tokens", "Best For", ""], HEADER_FONT, HEADER_FILL); r += 1
write_row(ws, r, ["Gemini 2.5 Flash", "$0.15", "$0.60", "Intent, query expansion, simple responses", ""]); r += 1
write_row(ws, r, ["Gemini 2.5 Pro", "$1.25", "$10.00", "Deep guidance, complex emotional responses", ""]); r += 1

r += 1
write_sub(ws, r, "Token Usage per Conversation Turn", COLS); r += 1
write_row(ws, r, ["API Call", "Model", "Input Tokens", "Output Tokens", "Calls per Turn"], HEADER_FONT, HEADER_FILL); r += 1
token_usage = [
    ["Intent classification", "Flash", "~300", "~80", "1"],
    ["Query expansion", "Flash", "~200", "~50", "1 (short queries only)"],
    ["HyDE generation", "Flash", "~200", "~100", "1 (if enabled)"],
    ["Response generation", "Flash or Pro", "~1,500", "~150", "1"],
]
for row_data in token_usage:
    write_row(ws, r, row_data); r += 1

r += 1
write_sub(ws, r, "Cost per Conversation (6 Turns Average)", COLS); r += 1
write_row(ws, r, ["Scenario", "Input Tokens", "Output Tokens", "Cost per Convo", "Notes"], HEADER_FONT, HEADER_FILL); r += 1
convo_cost = [
    ["Flash only", "~13,000", "~2,300", "$0.003", "Most cost-effective"],
    ["Flash + Pro (80/20 split)", "~13,000", "~2,300", "$0.011", "Better quality for guidance"],
    ["Pro only", "~13,000", "~2,300", "$0.039", "Highest quality, highest cost"],
]
for row_data in convo_cost:
    write_row(ws, r, row_data); r += 1

r += 1
write_sub(ws, r, "Monthly Cost by Volume", COLS); r += 1
write_row(ws, r, ["Conversations/Month", "Flash Only", "Flash+Pro (80/20)", "Pro Only", ""], HEADER_FONT, HEADER_FILL); r += 1
volume_cost = [
    ["1,000", "$3", "$11", "$39", ""],
    ["5,000", "$15", "$55", "$195", ""],
    ["10,000", "$30", "$110", "$390", ""],
    ["25,000", "$75", "$275", "$975", ""],
    ["50,000", "$150", "$550", "$1,950", ""],
    ["100,000", "$300", "$1,100", "$3,900", ""],
]
for row_data in volume_cost:
    write_row(ws, r, row_data); r += 1

r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
ws.cell(row=r, column=1, value="Note: Context caching (6h TTL) saves ~60-70% on input costs. System instruction (~2K tokens) is cached across all requests.").font = NOTE_FONT

auto_width(ws, COLS)

# ══════════════════════════════════════════════════════════════
# SHEET 6: GCP vs AWS Comparison
# ══════════════════════════════════════════════════════════════
ws = wb.create_sheet("GCP vs AWS Comparison")
COLS = 7
r = 1

write_title(ws, r, "GCP vs AWS — Side-by-Side Production Cost Comparison", COLS); r += 2

write_sub(ws, r, "Service Mapping", COLS); r += 1
write_row(ws, r, ["Component", "GCP Service", "AWS Equivalent", "GCP Low", "GCP High", "AWS Low", "AWS High"], HEADER_FONT, HEADER_FILL); r += 1

comparison = [
    ["Backend Compute", "Cloud Run (8vCPU/32GB)", "Fargate ECS (8vCPU/32GB)", "$715", "$858", "$340", "$408"],
    ["Load Balancer", "Built-in (Cloud Run)", "Application Load Balancer", "$0", "$0", "$21", "$26"],
    ["Frontend", "Vercel Pro", "Vercel Pro", "$20", "$20", "$20", "$20"],
    ["Database", "MongoDB Atlas M10", "MongoDB Atlas M10", "$57", "$100", "$57", "$100"],
    ["Cache (Redis)", "Memorystore 1GB Basic", "ElastiCache t4g.small", "$35", "$70", "$20", "$40"],
    ["LLM API", "Gemini Flash", "Gemini Flash", "$30", "$50", "$30", "$50"],
    ["Networking / VPC", "VPC Connector (e2-micro x2)", "NAT Gateway", "$13", "$74", "$41", "$92"],
    ["Storage / Registry", "Artifact Registry + GCS", "ECR + S3", "$1", "$8", "$1", "$6"],
    ["Monitoring", "Cloud Monitoring", "CloudWatch", "$0", "$17", "$0", "$20"],
    ["Backup / DR", "Included", "S3 export", "$0", "$2", "$0", "$2"],
    ["CI/CD", "Cloud Build", "CodeBuild", "$0", "$5", "$0", "$5"],
    ["Secrets", "Secret Manager (free)", "Secrets Manager", "$0", "$0", "$2", "$2"],
]
for row_data in comparison:
    write_row(ws, r, row_data); r += 1

write_row(ws, r, ["MONTHLY TOTAL", "", "", "$871", "$1,204", "$532", "$771"], TOTAL_FONT, TOTAL_FILL); r += 1
write_row(ws, r, ["ANNUAL TOTAL", "", "", "$10,452", "$14,448", "$6,384", "$9,252"], BOLD, GREEN_FILL); r += 1

r += 2
write_sub(ws, r, "Key Differences", COLS); r += 1
write_row(ws, r, ["Area", "GCP Advantage", "AWS Advantage", "", "", "", ""], HEADER_FONT, HEADER_FILL); r += 1
diffs = [
    ["Compute", "Cloud Run includes built-in LB, simpler setup", "Fargate is ~37% cheaper per vCPU/GB", "", "", "", ""],
    ["Redis", "Memorystore is well-integrated with Cloud Run", "ElastiCache is ~43% cheaper ($20 vs $36)", "", "", "", ""],
    ["Networking", "VPC Connector is cheap ($12/mo)", "NAT Gateway is more capable but pricier ($37/mo)", "", "", "", ""],
    ["Load Balancing", "Free (built into Cloud Run)", "ALB required ($21-26/mo additional)", "", "", "", ""],
    ["Secrets", "Free tier covers small usage", "Always costs $0.40/secret ($1.60/mo)", "", "", "", ""],
    ["Overall", "Simpler to manage, fewer moving parts", "~30% cheaper for same workload", "", "", "", ""],
]
for row_data in diffs:
    write_row(ws, r, row_data); r += 1

r += 2
write_sub(ws, r, "Savings Summary", COLS); r += 1
write_row(ws, r, ["Metric", "GCP", "AWS", "Savings", "Savings %", "", ""], HEADER_FONT, HEADER_FILL); r += 1
savings = [
    ["Monthly (Low)", "$871", "$532", "$339", "39%", "", ""],
    ["Monthly (High)", "$1,204", "$771", "$433", "36%", "", ""],
    ["Annual (Low)", "$10,452", "$6,384", "$4,068", "39%", "", ""],
    ["Annual (High)", "$14,448", "$9,252", "$5,196", "36%", "", ""],
]
for row_data in savings:
    write_row(ws, r, row_data, BOLD, GREEN_FILL); r += 1

r += 2
write_sub(ws, r, "Scaling Milestones (AWS)", COLS); r += 1
write_row(ws, r, ["Daily Active Users", "Fargate Config", "MongoDB", "Redis", "Gemini API", "Est. Monthly (Low)", "Est. Monthly (High)"], HEADER_FONT, HEADER_FILL); r += 1
milestones = [
    ["< 500", "2 vCPU, 4 GB, desired=0-1", "M0 (free)", "ElastiCache free tier", "Flash only", "$18", "$60"],
    ["500 - 2,000", "4 vCPU, 8 GB, desired=1, max=4", "M2 ($9)", "cache.t4g.small", "Flash only", "$180", "$380"],
    ["2,000 - 10,000 (CURRENT)", "8 vCPU, 32 GB, desired=1, max=100", "M10 ($57)", "cache.t4g.small", "Flash only", "$532", "$771"],
    ["10,000 - 50,000", "8 vCPU, 32 GB, desired=2, max=100", "M30 ($500)", "cache.r6g.large", "Flash + Pro", "$1,200", "$2,200"],
    ["50,000+", "8 vCPU, 32 GB, desired=5, max=100+", "M50+", "ElastiCache HA", "Flash + Pro", "$3,000", "$5,000+"],
]
for row_data in milestones:
    write_row(ws, r, row_data); r += 1

auto_width(ws, COLS)

# ══════════════════════════════════════════════════════════════
# SHEET 7: Cost Optimization
# ══════════════════════════════════════════════════════════════
ws = wb.create_sheet("Cost Optimization")
COLS = 4
r = 1

write_title(ws, r, "AWS Cost Optimization Recommendations", COLS); r += 2

write_row(ws, r, ["Optimization", "Monthly Savings", "Trade-off", "Priority"], HEADER_FONT, HEADER_FILL); r += 1
optimizations = [
    ["Use Graviton (ARM) Fargate tasks", "~$68/mo (20% on compute)", "Must verify Docker image builds for ARM; Python + ML libs have good ARM support", "High"],
    ["Fargate Spot for burst instances", "~$20-40/mo (70% on burst)", "Tasks can be interrupted during scaling; only for burst, not min instance", "Medium"],
    ["ECS Compute Savings Plan (1yr)", "~$100/mo (30% on compute)", "1-year commitment required; predictable workload makes this safe", "High"],
    ["Downsize to 4 vCPU / 8 GB", "~$170/mo", "May hit memory limits under heavy load; reranker/SPLADE already disabled", "High"],
    ["Set desired=0 (if low traffic)", "~$340/mo", "15-20s cold start on first request after idle", "High (if low traffic)"],
    ["Use Gemini Flash only (already done)", "Already active", "Pro not enabled currently", "Done"],
    ["Replace NAT Gateway with VPC endpoints", "~$30/mo", "Only saves if all traffic is to AWS services (S3, ECR, etc.)", "Medium"],
    ["Use S3 VPC endpoint (free)", "~$2-5/mo on data transfer", "Gateway endpoint is free; no change to functionality", "Low"],
    ["ElastiCache Reserved Instance (1yr)", "~$7/mo (35% on Redis)", "1-year commitment", "Low"],
    ["Gemini context caching (already enabled)", "~60-70% on input tokens", "None (already active)", "Done"],
    ["Semantic response cache (already enabled)", "Saves 5-15s on repeat queries", "None (already active)", "Done"],
]
for row_data in optimizations:
    write_row(ws, r, row_data); r += 1

r += 2
write_sub(ws, r, "Potential Savings Summary", COLS); r += 1
write_row(ws, r, ["Scenario", "Current Monthly", "Optimized Monthly", "Annual Savings"], HEADER_FONT, HEADER_FILL); r += 1
opt_savings = [
    ["Graviton + Savings Plan (no downsize)", "$532 - $771", "$362 - $601", "$2,040 - $2,040"],
    ["Graviton + Savings Plan + 4vCPU/8GB", "$532 - $771", "$192 - $431", "$4,080 - $4,080"],
    ["Scale-to-zero (low traffic periods)", "$532 - $771", "$192 - $431", "$4,080 - $4,080"],
]
for row_data in opt_savings:
    write_row(ws, r, row_data, BOLD, GREEN_FILL); r += 1

auto_width(ws, COLS)

# ══════════════════════════════════════════════════════════════
# SHEET 8: AWS Cost Calculator (detailed — mirrors GCP COST CAL)
# ══════════════════════════════════════════════════════════════
ws = wb.create_sheet("AWS Cost Calculator")
COLS = 7
r = 1

write_title(ws, r, "3ioNetra — AWS Cost Calculator (Production)", COLS); r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
ws.cell(row=r, column=1, value="Exact pricing from AWS rate cards. Region: eu-west-1 (Ireland). All rates as of March 2026.").font = ITALIC_GRAY; r += 2

# AWS Pricing Rates Reference
write_sub(ws, r, "AWS Pricing Rates (Reference)", COLS); r += 1
write_row(ws, r, ["Service", "SKU", "Rate", "Unit", "Region / Tier", "", ""], HEADER_FONT, HEADER_FILL); r += 1

aws_rates = [
    ["Fargate", "vCPU (Linux/x86)", "$0.04048", "per vCPU-hour", "eu-west-1", "= $0.000011244/vCPU-sec", ""],
    ["Fargate", "Memory (Linux/x86)", "$0.004445", "per GB-hour", "eu-west-1", "= $0.000001235/GB-sec", ""],
    ["Fargate", "vCPU (Linux/ARM Graviton)", "$0.03238", "per vCPU-hour", "eu-west-1", "20% savings", ""],
    ["Fargate", "Memory (Linux/ARM Graviton)", "$0.003560", "per GB-hour", "eu-west-1", "20% savings", ""],
    ["Fargate", "Ephemeral storage", "Free", "20 GB included per task", "All regions", "", ""],
    ["ElastiCache", "cache.t4g.small (1.37 GiB)", "$0.028", "per hour", "eu-west-1", "", ""],
    ["ElastiCache", "cache.t4g.micro (0.5 GiB)", "$0.018", "per hour", "eu-west-1", "", ""],
    ["ALB", "Hourly charge", "$0.0225", "per hour", "eu-west-1", "", ""],
    ["ALB", "LCU charge", "$0.008", "per LCU-hour", "eu-west-1", "", ""],
    ["NAT Gateway", "Hourly charge", "$0.048", "per hour", "eu-west-1", "", ""],
    ["NAT Gateway", "Data processing", "$0.048", "per GB", "eu-west-1", "", ""],
    ["ECR", "Storage", "$0.10", "per GB per month", "All regions", "", ""],
    ["S3", "Standard storage", "$0.023", "per GB per month", "eu-west-1", "", ""],
    ["CodeBuild", "general1.small", "$0.005", "per build-minute", "After 100 free min/mo", "", ""],
    ["CloudWatch", "Log ingestion", "$0.50", "per GB", "After 5 GB free/mo", "", ""],
    ["Secrets Manager", "Per secret", "$0.40", "per secret per month", "All regions", "", ""],
]
for row_data in aws_rates:
    write_row(ws, r, row_data, NORMAL, RATE_FILL); r += 1

r += 1

# Resource-by-Resource Calculation
write_sub(ws, r, "Resource-by-Resource Cost Calculation", COLS); r += 1
write_row(ws, r, ["Resource", "SKU / Config", "Quantity", "Unit", "Rate ($/unit)", "Calculation", "Monthly Cost ($)"], HEADER_FONT, HEADER_FILL); r += 1

# Fargate min instance
write_row(ws, r, ["FARGATE ECS — backend (desired=1, 8 vCPU, 32 GB)", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
fargate_rows = [
    ["  Min task vCPU", "vCPU always allocated", "8 vCPU x 730 hrs", "vCPU-hours", "$0.04048", "8 x 730 x $0.04048", "$236.40"],
    ["  Min task Memory", "Memory always allocated", "32 GB x 730 hrs", "GB-hours", "$0.004445", "32 x 730 x $0.004445", "$103.82"],
]
for row_data in fargate_rows:
    write_row(ws, r, row_data); r += 1
write_row(ws, r, ["  FARGATE SUBTOTAL (min task only)", "", "", "", "", "", "$340.22"], BOLD, YELLOW_FILL); r += 1

r += 1
write_row(ws, r, ["FARGATE ECS — Burst (additional tasks at peak)", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
burst_rows = [
    ["  Burst vCPU (avg 1 extra task, 20% time)", "vCPU always allocated", "8 vCPU x 146 hrs", "vCPU-hours", "$0.04048", "8 x 146 x $0.04048", "$47.28"],
    ["  Burst Memory", "Memory always allocated", "32 GB x 146 hrs", "GB-hours", "$0.004445", "32 x 146 x $0.004445", "$20.76"],
]
for row_data in burst_rows:
    write_row(ws, r, row_data); r += 1
write_row(ws, r, ["  FARGATE BURST SUBTOTAL", "", "", "", "", "", "$68.04"], BOLD, YELLOW_FILL); r += 1

r += 1
# ALB
write_row(ws, r, ["APPLICATION LOAD BALANCER", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
alb_calc = [
    ["  ALB hourly charge", "Always on", "730 hrs", "hours", "$0.0225", "730 x $0.0225", "$16.43"],
    ["  LCU charges (estimated)", "~5-10 LCU-hours avg", "~7 LCU-hrs avg", "LCU-hours", "$0.008", "7 x 730 x $0.008", "$8.00"],
]
for row_data in alb_calc:
    write_row(ws, r, row_data); r += 1
write_row(ws, r, ["  ALB SUBTOTAL", "", "", "", "", "", "$24.43"], BOLD, YELLOW_FILL); r += 1

r += 1
# NAT Gateway
write_row(ws, r, ["NAT GATEWAY (private subnet access)", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
nat_calc = [
    ["  NAT Gateway hourly", "Always on", "730 hrs", "hours", "$0.048", "730 x $0.048", "$35.04"],
    ["  NAT Gateway data processing", "~50 GB/month", "50 GB", "GB", "$0.048", "50 x $0.048", "$2.40"],
]
for row_data in nat_calc:
    write_row(ws, r, row_data); r += 1
write_row(ws, r, ["  NAT GATEWAY SUBTOTAL", "", "", "", "", "", "$37.44"], BOLD, YELLOW_FILL); r += 1

r += 1
# ElastiCache
write_row(ws, r, ["ELASTICACHE REDIS — cache.t4g.small (1.37 GiB)", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
write_row(ws, r, ["  cache.t4g.small On-Demand", "1 node, 730 hrs/mo", "1 x 730 hrs", "hours", "$0.028", "1 x 730 x $0.028", "$20.44"]); r += 1
write_row(ws, r, ["  ELASTICACHE SUBTOTAL", "", "", "", "", "", "$20.44"], BOLD, YELLOW_FILL); r += 1

r += 1
# ECR
write_row(ws, r, ["AMAZON ECR — Docker image storage", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
write_row(ws, r, ["  Docker image storage", "~6.2 GB", "6.2 GB", "GB-months", "$0.10", "6.2 x $0.10", "$0.62"]); r += 1
write_row(ws, r, ["  ECR SUBTOTAL", "", "", "", "", "", "$0.62"], BOLD, YELLOW_FILL); r += 1

r += 1
# S3
write_row(ws, r, ["AMAZON S3 (3ioNetra buckets only)", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
s3_rows = [
    ["  CodeBuild artifacts", "eu-west-1, 1.5 GB", "1.5 GB", "GB-months", "$0.023", "1.5 x $0.023", "$0.03"],
    ["  Source uploads", "eu-west-1, 0.5 GB", "0.5 GB", "GB-months", "$0.023", "0.5 x $0.023", "$0.01"],
]
for row_data in s3_rows:
    write_row(ws, r, row_data); r += 1
write_row(ws, r, ["  S3 SUBTOTAL", "", "", "", "", "", "$0.04"], BOLD, YELLOW_FILL); r += 1

r += 1
# CodeBuild
write_row(ws, r, ["AWS CODEBUILD", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
write_row(ws, r, ["  Build minutes", "100 free min/month", "~100 min/mo used", "build-minutes", "$0.005", "Within free tier", "$0.00"]); r += 1
write_row(ws, r, ["  CODEBUILD SUBTOTAL", "", "", "", "", "", "$0.00"], BOLD, YELLOW_FILL); r += 1

r += 1
# CloudWatch
write_row(ws, r, ["AMAZON CLOUDWATCH", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
write_row(ws, r, ["  Log ingestion", "5 GB free/mo", "~5-10 GB/mo used", "GB", "$0.50", "Within free tier", "$0.00"]); r += 1
write_row(ws, r, ["  CLOUDWATCH SUBTOTAL", "", "", "", "", "", "$0.00"], BOLD, YELLOW_FILL); r += 1

r += 1
# Secrets Manager
write_row(ws, r, ["AWS SECRETS MANAGER", "", "", "", "", "", ""], BOLD, SUB_FILL); r += 1
write_row(ws, r, ["  4 secrets", "Gemini, MongoDB, Redis, DB password", "4 secrets", "secrets/mo", "$0.40", "4 x $0.40", "$1.60"]); r += 1
write_row(ws, r, ["  SECRETS MANAGER SUBTOTAL", "", "", "", "", "", "$1.60"], BOLD, YELLOW_FILL); r += 1

r += 2

# AWS Cost Summary
write_sub(ws, r, "AWS Cost Summary", COLS); r += 1
write_row(ws, r, ["Component", "", "", "", "", "Low ($/mo)", "High ($/mo)"], HEADER_FONT, HEADER_FILL); r += 1

aws_summary = [
    ["Fargate ECS — min task (always on)", "", "", "", "", "$340.22", "$340.22"],
    ["Fargate ECS — burst (avg 1 extra, 20% time)", "", "", "", "", "$0.00", "$68.04"],
    ["Application Load Balancer", "", "", "", "", "$24.43", "$24.43"],
    ["NAT Gateway (private subnet)", "", "", "", "", "$37.44", "$37.44"],
    ["ElastiCache Redis (cache.t4g.small)", "", "", "", "", "$20.44", "$20.44"],
    ["Amazon ECR (6.2 GB)", "", "", "", "", "$0.62", "$0.62"],
    ["Amazon S3 (2 buckets, ~2 GB)", "", "", "", "", "$0.04", "$0.04"],
    ["AWS CodeBuild", "", "", "", "", "$0.00", "$0.00"],
    ["Amazon CloudWatch", "", "", "", "", "$0.00", "$0.00"],
    ["AWS Secrets Manager", "", "", "", "", "$1.60", "$1.60"],
]
for row_data in aws_summary:
    write_row(ws, r, row_data); r += 1

write_row(ws, r, ["AWS SUBTOTAL", "", "", "", "", "$424.79", "$492.83"], TOTAL_FONT, TOTAL_FILL); r += 1

r += 2

# External Services
write_sub(ws, r, "External Services (Not AWS)", COLS); r += 1
write_row(ws, r, ["Service", "Provider", "Tier / Plan", "Details", "", "Low ($/mo)", "High ($/mo)"], HEADER_FONT, HEADER_FILL); r += 1

ext = [
    ["MongoDB Atlas", "MongoDB Inc.", "M10 dedicated", "2 vCPU, 2 GB RAM, 10 GB storage", "", "$57.00", "$57.00"],
    ["MongoDB Atlas (scale-up)", "MongoDB Inc.", "M30 (if needed)", "For 50K+ users", "", "$0.00", "$443.00"],
    ["Vercel (Frontend)", "Vercel", "Pro plan", "Next.js hosting, 1 TB bandwidth", "", "$20.00", "$20.00"],
    ["Gemini API", "Google", "Flash only (~10K convos)", "$0.003/convo x 10K", "", "$30.00", "$50.00"],
    ["Domain + DNS", "Cloudflare", "Free plan + domain renewal", "3iosetu.com, auto SSL", "", "$1.00", "$1.00"],
    ["Cloudflare Pro (optional)", "Cloudflare", "Pro plan", "WAF, advanced DDoS", "", "$0.00", "$20.00"],
]
for row_data in ext:
    write_row(ws, r, row_data); r += 1

write_row(ws, r, ["EXTERNAL SUBTOTAL", "", "", "", "", "$108.00", "$591.00"], TOTAL_FONT, TOTAL_FILL); r += 1

r += 2

# Grand Total
write_sub(ws, r, "GRAND TOTAL — 3ioNetra Production Infrastructure (AWS)", COLS); r += 1
write_row(ws, r, ["", "", "", "", "", "Low ($/mo)", "High ($/mo)"], HEADER_FONT, HEADER_FILL); r += 1
write_row(ws, r, ["AWS Services", "", "", "", "", "$424.79", "$492.83"], BOLD); r += 1
write_row(ws, r, ["External Services", "", "", "", "", "$108.00", "$591.00"], BOLD); r += 1
r += 1
write_row(ws, r, ["TOTAL MONTHLY", "", "", "", "", "$532.79", "$1,083.83"], TOTAL_FONT, TOTAL_FILL); r += 1
write_row(ws, r, ["TOTAL ANNUAL", "", "", "", "", "$6,393.48", "$13,005.96"], BOLD, GREEN_FILL); r += 1

r += 2

# Comparison with GCP
write_sub(ws, r, "vs GCP Comparison", COLS); r += 1
write_row(ws, r, ["", "", "", "", "", "Low ($/mo)", "High ($/mo)"], HEADER_FONT, HEADER_FILL); r += 1
write_row(ws, r, ["GCP Total Monthly", "", "", "", "", "$699.36", "$1,291.69"], BOLD); r += 1
write_row(ws, r, ["AWS Total Monthly", "", "", "", "", "$532.79", "$1,083.83"], BOLD); r += 1
write_row(ws, r, ["MONTHLY SAVINGS (AWS)", "", "", "", "", "$166.57", "$207.86"], BOLD, GREEN_FILL); r += 1
write_row(ws, r, ["ANNUAL SAVINGS (AWS)", "", "", "", "", "$1,998.84", "$2,494.32"], BOLD, GREEN_FILL); r += 1
write_row(ws, r, ["SAVINGS %", "", "", "", "", "24%", "16%"], BOLD, GREEN_FILL); r += 1

r += 2

# Notes
write_sub(ws, r, "Notes & Assumptions", COLS); r += 1
notes = [
    "1. Fargate pricing is for Linux/x86 in eu-west-1. ARM/Graviton offers 20% savings if Docker image supports ARM.",
    "2. Burst estimate assumes avg 1 additional task running 20% of the time (peak hours). Actual varies with traffic.",
    "3. ALB is required for Fargate (Cloud Run includes built-in load balancing at no extra cost).",
    "4. NAT Gateway is the equivalent of GCP VPC Connector for private subnet access (Redis, etc.).",
    "5. ElastiCache cache.t4g.small (1.37 GiB) is the closest match to GCP Memorystore Basic 1 GB.",
    "6. AWS Secrets Manager charges $0.40/secret/month; GCP Secret Manager free tier covers small usage.",
    "7. MongoDB Atlas pricing is identical on AWS and GCP. Gemini API is cloud-agnostic.",
    "8. Fargate has no built-in free tier (unlike Cloud Run). Free tier for ElastiCache is 750 hrs/12 mo for new accounts.",
    "9. Pricing sourced from AWS rate cards and Vantage.sh as of March 2026.",
]
r = write_notes(ws, r, notes, COLS)

auto_width(ws, COLS)

# ── Save ──
OUTPUT = "/Users/ankit1609/Desktop/AWS_Costing.xlsx"
wb.save(OUTPUT)
print(f"Generated: {OUTPUT}")
print(f"\nSheets: {wb.sheetnames}")
print(f"\nAWS Monthly: $532.79 - $1,083.83")
print(f"GCP Monthly: $699.36 - $1,291.69")
print(f"Savings:     $166.57 - $207.86/mo (16-24%)")
